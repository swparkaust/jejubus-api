import datetime
import difflib
import os
import re
import sys
import inquirer
import requests
import requests_cache
import xmltodict
from openpyxl import load_workbook
from tqdm import tqdm

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db.models import Func, F

from main.models import Route, Station, StationOtherName, StationRoute, Time


def download(url: str, dest_folder: str):
    if not os.path.exists(dest_folder):
        os.makedirs(dest_folder)  # create folder if it does not exist

    filename = url.split('/')[-1].replace(
        " ", "_") + ".xlsx"  # be careful with file names
    file_path = os.path.join(dest_folder, filename)

    r = requests.get(url, stream=True)
    if r.ok:
        print("saving to", os.path.abspath(file_path))
        with open(file_path, 'wb') as f:
            for chunk in r.iter_content(chunk_size=1024 * 8):
                if chunk:
                    f.write(chunk)
                    f.flush()
                    os.fsync(f.fileno())
    else:  # HTTP status code 4XX/5XX
        print("Download failed: status code {}\n{}".format(r.status_code, r.text))


def extract_holiday_types_from_string(s):
    pattern = r'(평|토(?:요)?|공휴|주말)(?:일)?'
    matches = re.findall(pattern, s)
    if not matches:
        return ["1", "2", "3"]
    else:
        holiday_types = []
        for match in matches:
            if match == "평":
                holiday_types.append("1")
            elif match == "토":
                holiday_types.append("2")
            elif match == "공휴":
                holiday_types.append("3")
            elif match == "주말":
                holiday_types.append("2")
                holiday_types.append("3")
        return holiday_types


def extract_node_name_from_string(s):
    pattern = r'(.*?)(?:\(.*\))?(?:\[.*\])?$'
    matches = re.findall(pattern, s)
    if not matches:
        return s
    else:
        return matches[0]


def extract_route_number_from_string(s):
    pattern = r'\d[0-9\-]*'
    matches = re.findall(pattern, s)
    if not matches:
        return s.replace('\n', '').replace('\r', '')
    else:
        return matches[0]


def extract_time_from_string(s):
    pattern = r'(2[0-3]|[01]?[0-9])(?::|;)([0-5]?[0-9])'
    matches = re.findall(pattern, s)
    if not matches:
        return None
    else:
        return datetime.time(int(matches[0][0]), int(matches[0][1]), 0)


def get_all_routes():
    url = 'http://busopen.jeju.go.kr/OpenAPI/service/bis/Bus'

    r = requests.get(url)
    data = xmltodict.parse(r.content)

    return data['response']['body']['items']['item']


def get_all_station_routes():
    url = 'http://busopen.jeju.go.kr/OpenAPI/service/bis/StationRoute'

    r = requests.get(url)
    data = xmltodict.parse(r.content)

    return data['response']['body']['items']['item']


def get_all_stations():
    url = 'http://busopen.jeju.go.kr/OpenAPI/service/bis/Station'

    r = requests.get(url)
    data = xmltodict.parse(r.content)

    return data['response']['body']['items']['item']


def get_first_cell(worksheet):
    for row_cells in worksheet.iter_rows():
        for cell in row_cells:
            if cell.value is not None:
                return cell


def get_node_ids(node_name):
    return [x.station_id for x in Station.objects.filter(station_name__icontains=node_name)]


def get_route_node(route_id, node_name, start=None, end=None, interactive=True):
    sl = slice(start, end)
    station_routes = list(StationRoute.objects.filter(
        route__route_id=route_id).order_by('station_order'))[sl]
    node_ids = get_node_ids(node_name)
    for station_route in station_routes:
        for node_id in node_ids:
            if station_route.station.station_id == node_id:
                return station_route
    station_other_names = StationOtherName.objects.filter(
        other_station_name=node_name)
    if station_other_names.exists():
        for station_route in station_routes:
            for o in station_other_names:
                if station_route.station.station_id == o.station_id:
                    return station_route
    if interactive:
        choices = [(station_route.station.station_name, i)
                   for i, station_route in enumerate(station_routes)]
        if choices:
            choices.sort(key=lambda x: difflib.SequenceMatcher(
                None, x[0], node_name).ratio(), reverse=True)
            questions = [
                inquirer.List(
                    'node_name',
                    message="What node is " + node_name + "?",
                    choices=choices, ),
            ]
            answers = inquirer.prompt(questions)
            if answers is None:
                return None
            else:
                selected_node = station_routes[answers["node_name"]]
                station_other_name = StationOtherName(
                    station_id=selected_node.station.station_id, other_station_name=node_name)
                station_other_name.save()
                return selected_node
        else:
            return None
    else:
        return None


def get_route_nodes(route_id, node_names):
    route_nodes = []
    last = None
    for i, node_name in enumerate(node_names):
        if last is None:
            route_node = get_route_node(route_id, node_name, 0, 1, False)
        else:
            route_node = get_route_node(
                route_id, node_name, last, -(len(node_names) - i - 1) or None, False)
        if route_node is None:
            continue
        route_nodes.append(route_node)
        last = route_node.station_order
    return route_nodes


def get_route(route_number, node_names, interactive=True):
    routes = Route.objects.filter(route_number__icontains=route_number)
    if routes:
        route_nodes_list = [get_route_nodes(
            route.route_id, node_names) for route in routes]
        choices = [("-".join([x.station.station_name for x in route_nodes]), i)
                   for i, route_nodes in enumerate(route_nodes_list)]
        if choices:
            choices.sort(key=lambda x: len(
                route_nodes_list[x[1]]), reverse=True)
            if interactive:
                questions = [
                    inquirer.List(
                        'route',
                        message="What route is " + route_number +
                        " " + "-".join(node_names) + "?",
                        choices=choices, ),
                ]
                answers = inquirer.prompt(questions)
                if answers is None:
                    return None
                else:
                    selected_route = routes[answers["route"]]
            else:
                selected_route = routes[choices[0][1]]
            station_routes = StationRoute.objects.filter(
                route__route_id=selected_route.route_id).order_by('station_order')
            start_station = station_routes.first().station
            end_station = station_routes.last().station
            if start_station.station_name != node_names[0] and not StationOtherName.objects.filter(other_station_name=node_names[0]).exists():
                start_station_other_name = StationOtherName(
                    station_id=start_station.station_id, other_station_name=node_names[0])
                start_station_other_name.save()
            if end_station.station_name != node_names[-1] and not StationOtherName.objects.filter(other_station_name=node_names[-1]).exists():
                end_station_other_name = StationOtherName(
                    station_id=end_station.station_id, other_station_name=node_names[-1])
                end_station_other_name.save()
            return selected_route
        else:
            return None
    else:
        return None


class Command(BaseCommand):

    help = 'Updates the database via Bus Info API'

    def add_arguments(self, parser):
        parser.add_argument(
            '--clear-history',
            action='store_true',
            dest='clear_history',
            help='Clean station find history',
        )
        parser.add_argument(
            '--clear-db',
            action='store_true',
            dest='clear_db',
            help='Clear database before adding new objects',
        )
        parser.add_argument(
            '--noinput',
            action='store_false',
            dest='interactive',
            default=True,
            help='Do NOT prompt the user for input of any kind.',
        )

    def handle(self, *args, **options):
        requests_cache.install_cache('jejubus_cache')

        if options['clear_history']:
            sys.stdout.write('Clearing station find history ... ')
            sys.stdout.flush()
            StationOtherName.objects.all().delete()
            sys.stdout.write('done.\n')

        if options['clear_db']:
            sys.stdout.write('Clearing database ... ')
            sys.stdout.flush()
            Time.objects.all().delete()
            StationRoute.objects.all().delete()
            Route.objects.all().delete()
            Station.objects.all().delete()
            sys.stdout.write('done.\n')

        for route_type in settings.ROUTE_TYPES:
            download(
                "http://bus.jeju.go.kr/publicTrafficInformation/downloadSchedule/" +
                route_type,
                dest_folder="temp")

        sys.stdout.write('Saving routes ... ')
        sys.stdout.flush()
        for route in get_all_routes():
            route_obj = Route(
                route_type=route['routeTp'], route_id=route['routeId'], route_number=route['routeNum'])
            route_obj.save()
        sys.stdout.write('done.\n')

        sys.stdout.write('Saving stations ... ')
        sys.stdout.flush()
        for station in get_all_stations():
            station_obj = Station(
                local_x=station['localX'], local_y=station['localY'], station_id=station['stationId'], station_name=station['stationNm'])
            station_obj.save()
        sys.stdout.write('done.\n')

        sys.stdout.write('Saving station routes ... ')
        sys.stdout.flush()
        for station_route in get_all_station_routes():
            route = Route.objects.get(route_id=station_route['routeId'])
            station = Station.objects.get(
                station_id=station_route['stationId'])
            station_route_obj = StationRoute(route=route, station=station, station_order=int(
                station_route['stationOrd']), up_down_direction=station_route['updnDir'])
            station_route_obj.save()
        sys.stdout.write('done.\n')

        with os.scandir("temp") as it:
            items = (entry for entry in it if entry.name.endswith(
                ".xlsx") and entry.is_file())

            pbar = tqdm(items, total=len(os.listdir("temp")))
            for entry in pbar:
                pbar.set_description("Processing %s" % entry.name)

                wb = load_workbook(filename=entry.path, data_only=True)

                pbar2 = tqdm(wb.worksheets)
                for sheet in pbar2:
                    first_cell = get_first_cell(sheet)

                    route_number = first_cell.value
                    holiday_types = extract_holiday_types_from_string(
                        route_number)
                    route_number = extract_route_number_from_string(
                        route_number)

                    node_names = []
                    row = first_cell.offset(row=5).row
                    for cell in sheet[row]:
                        node_name = cell.value
                        if node_name is not None:
                            node_name = "".join(node_name.split())
                            if node_name != "노선번호" and node_name != "구분" and node_name != "비고":
                                node_name = extract_node_name_from_string(
                                    node_name)
                                node_names.append(node_name)

                    route = get_route(
                        route_number, node_names, options['interactive'])
                    if route is None:
                        continue

                    pbar2.set_description(
                        "Processing route %s" % route.route_number)

                    route_number_column = None
                    last = None
                    i = 0
                    for cell in sheet[row]:
                        node_name = cell.value
                        if node_name is not None:
                            node_name = "".join(node_name.split())
                            if node_name == "노선번호":
                                route_number_column = cell.column_letter
                            elif node_name != "구분" and node_name != "비고":
                                node_name = extract_node_name_from_string(
                                    node_name)
                                if last is None:
                                    route_node = get_route_node(
                                        route.route_id, node_name, 0, 1, options['interactive'])
                                else:
                                    route_node = get_route_node(
                                        route.route_id, node_name, last, -(len(node_names) - i - 1) or None, options['interactive'])
                                if route_node is None:
                                    continue
                                last = route_node.station_order
                                i += 1
                                station = route_node.station
                                node_name = station.station_name
                                for cell2 in sheet[cell.column][cell.row + 1:]:
                                    time = cell2.value
                                    if time is not None:
                                        if isinstance(time, str):
                                            time = extract_time_from_string(
                                                time)
                                            if time is None:
                                                continue
                                        elif type(time) is not datetime.time:
                                            continue
                                        if route_number_column is not None:
                                            cell_name = "{}{}".format(
                                                route_number_column, cell2.row)
                                            route_number = sheet[cell_name].value
                                            route_number = extract_route_number_from_string(
                                                str(route_number))

                                            route = get_route(
                                                route_number, node_names, options['interactive'])
                                            if route is None:
                                                continue
                                        station_route = StationRoute.objects.filter(route=route, station=station).annotate(abs_diff=Func((F('station_order') - 1) / (StationRoute.objects.filter(
                                            route=route).count() - 1) - (route_node.station_order - 1) / (StationRoute.objects.filter(route=route_node.route).count() - 1), function='ABS')).order_by('abs_diff').first()
                                        if station_route:
                                            for holiday_type in holiday_types:
                                                time_obj = Time(
                                                    holiday_type=holiday_type, station_route=station_route, time=time)
                                                time_obj.save()

        sys.stdout.write(self.style.SUCCESS(
            'Successfully updated the database\n'))
